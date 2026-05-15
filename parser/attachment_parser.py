import io
import tempfile
import os
from pathlib import Path

import httpx
import pdfplumber
import openpyxl


async def download_file(url: str) -> bytes | None:
    """Download a file from URL. Returns bytes or None on failure."""
    try:
        async with httpx.AsyncClient(verify=False, follow_redirects=True, timeout=60) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.content
    except Exception as e:
        print(f"下载失败 {url}: {e}")
        return None


def extract_text_from_pdf(content: bytes) -> str:
    """Extract text from PDF bytes using pdfplumber."""
    try:
        text_parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages[:20]:  # Limit to first 20 pages
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
                # Also try extracting tables
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if row:
                            text_parts.append(" | ".join(str(cell or "") for cell in row))
        return "\n".join(text_parts)
    except Exception as e:
        print(f"PDF解析失败: {e}")
        return ""


def extract_text_from_excel(content: bytes) -> str:
    """Extract text from Excel bytes using openpyxl."""
    try:
        text_parts = []
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames[:5]:  # Limit to first 5 sheets
            ws = wb[sheet_name]
            text_parts.append(f"=== Sheet: {sheet_name} ===")
            row_count = 0
            for row in ws.iter_rows(max_row=200, values_only=True):  # Limit rows
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(cells):
                    text_parts.append(" | ".join(cells))
                    row_count += 1
            if row_count == 0:
                text_parts.append("(空表)")
        wb.close()
        return "\n".join(text_parts)
    except Exception as e:
        print(f"Excel解析失败: {e}")
        return ""


def get_file_type(url: str, content: bytes) -> str:
    """Detect file type from URL and content."""
    url_lower = url.lower()
    if url_lower.endswith(".pdf") or b"%PDF" in content[:10]:
        return "pdf"
    if url_lower.endswith((".xlsx", ".xls")) or b"xl/" in content[:100]:
        return "xlsx"
    if url_lower.endswith((".doc", ".docx")):
        return "docx"
    return "unknown"


# 附件相关性关键词（文件名或上下文包含以下词才下载）
ATTACHMENT_KEYWORDS = [
    "岗位", "招聘", "招录", "职位", "计划表", "人员需求",
    "录用", "聘用", "用人", "公招", "选调", "遴选",
    "岗位表", "职位表", "需求表", "名额",
]

# 排除关键词（这些附件通常不是岗位表）
ATTACHMENT_EXCLUDE = [
    "报名", "须知", "承诺书", "诚信", "协议", "合同",
    "体检", "政审", "考察", "公示", "通知", "公告",
    "指南", "样表", "模板", "填写说明",
]


def should_download(url: str, context: str = "") -> bool:
    """Decide whether an attachment is worth downloading.
    Checks filename and surrounding context for position-related keywords.
    """
    from urllib.parse import unquote
    filename = unquote(url.split("/")[-1].split("?")[0]).lower()
    context_lower = context.lower()

    # Check exclude keywords first
    for kw in ATTACHMENT_EXCLUDE:
        if kw in filename or kw in context_lower:
            return False

    # Check positive keywords
    for kw in ATTACHMENT_KEYWORDS:
        if kw in filename or kw in context_lower:
            return True

    # No keywords matched - skip
    return False


async def process_attachment(url: str) -> str:
    """Download and extract text from an attachment.
    Returns extracted text or empty string.
    """
    content = await download_file(url)
    if not content:
        return ""

    file_type = get_file_type(url, content)
    print(f"    附件类型: {file_type}, 大小: {len(content)/1024:.1f}KB")

    if file_type == "pdf":
        return extract_text_from_pdf(content)
    elif file_type == "xlsx":
        return extract_text_from_excel(content)
    else:
        # Try PDF first, then Excel
        text = extract_text_from_pdf(content)
        if not text.strip():
            text = extract_text_from_excel(content)
        return text
